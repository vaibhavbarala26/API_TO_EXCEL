import os
from requests import Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
import json
import time
import pandas as pd
from openpyxl import load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from dotenv import load_dotenv
load_dotenv(dotenv_path="./env/.env")
# API Key for CoinMarketCap
api_key = os.getenv("api_key")

# Function to check if a file is locked (used for checking if Excel files are open)
def is_file_locked(file_path):
    """Check if a file is locked by trying to open it in write mode."""
    try:
        with open(file_path, 'a'):
            pass  # File can be opened without errors
        return False
    except IOError:
        return True  # File is locked

# Function to fetch cryptocurrency data and sort it
def fetch_add_to_excel(api_key, limit=50):
    """Fetch and sort cryptocurrency data from the CoinMarketCap API."""
    url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest"
    parameters = {
        "start": "1",  # Starting rank
        "limit": limit,  # Number of cryptocurrencies to fetch
        "convert": "USD",  # Convert prices to USD
    }
    headers = {
        "Accepts": "application/json",
        "X-CMC_PRO_API_KEY": api_key,  # API key header
    }

    session = Session()
    session.headers.update(headers)

    try:
        # Send a GET request to the API
        response = session.get(url, params=parameters)
        data = response.json()

        # Check for successful response
        if response.status_code != 200:
            print(f"API Error: {data.get('status', {}).get('error_message', 'Unknown error')}")
            return []

        # Parse the response data
        crypto_data = []
        for coin in data["data"]:
            crypto_data.append({
                "Name": coin["name"],
                "Symbol": coin["symbol"],
                "Price (USD)": coin["quote"]["USD"]["price"],
                "24h Trading Volume (USD)": coin["quote"]["USD"]["volume_24h"],
                "Market Cap (USD)": coin["quote"]["USD"]["market_cap"],
                "24h % Change": coin["quote"]["USD"]["percent_change_24h"]
            })

        # Sort the data by price in descending order
        return sorted(crypto_data, key=lambda x: x["Price (USD)"], reverse=True)

    except (ConnectionError, Timeout, TooManyRedirects) as e:
        print(f"An error occurred: {e}")
        return []

# Authenticate and create Google Drive service client using service account
def authenticate_google_drive(service_account_json):
    """Authenticate the service account and create the Google Drive service client."""
    SCOPES = ['https://www.googleapis.com/auth/drive']

    creds = service_account.Credentials.from_service_account_file(service_account_json, scopes=SCOPES)
    service = build('drive', 'v3', credentials=creds)
    return service

# Check if the file already exists in Google Drive folder
def file_exists(drive_service, file_name, parent_folder):
    """Check if a file with the given name exists in the specified Google Drive folder."""
    query = f"'{parent_folder}' in parents and name = '{file_name}'"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get('files', [])
    return files[0] if files else None

# Upload or update a file to Google Drive
def upload_or_update_file(file_path, drive_service, parent_folder):
    """Upload a new file to Google Drive or update if it exists."""
    file_name = os.path.basename(file_path)
    
    # Check if the file already exists
    existing_file = file_exists(drive_service, file_name, parent_folder)
    
    if existing_file:
        # File exists, update it
        file_id = existing_file['id']
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Update the file
        updated_file = drive_service.files().update(
            fileId=file_id,
            media_body=media
        ).execute()
        
        print(f"File updated successfully, File ID: {updated_file['id']}")
    else:
        # File does not exist, upload a new one
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_metadata = {
            'name': file_name,
            'parents': [parent_folder]  # Set the parent folder ID
        }
        
        new_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        print(f"File uploaded successfully, File ID: {new_file['id']}")

# Main execution logic
api_key = api_key
output_file = "crypto_data.xlsx"
percentage_change_file = "percentage_change.xlsx"
service_account_json = "./Credentials.json"  # Path to your service account credentials
parent_folder = os.getenv("parent_folder")# Google Drive folder ID

try:
    while True:
        # Fetch and sort cryptocurrency data
        crypto_data_sorted = fetch_add_to_excel(api_key, limit=50)

        if crypto_data_sorted:
            # Create a DataFrame from the fetched data
            df = pd.DataFrame(crypto_data_sorted)

            # Sort the data by 24h percentage change
            percentage_change_sorted = df.sort_values(by="24h % Change", ascending=False)

            # Calculate the average price of the top 50 cryptocurrencies
            average_price_of_50_cryptos = float(sum([coin["Price (USD)"] for coin in crypto_data_sorted]) / 50)

            # Check if any output files are locked
            if is_file_locked(output_file) or is_file_locked(percentage_change_file):
                print("One of the Excel files is currently open. Please close it to allow updates.")
            else:
                try:
                    # Save the main data to the output file
                    df.to_excel(output_file, index=False, engine="openpyxl")

                    # Save the percentage change sorted data to a separate file
                    percentage_change_sorted.to_excel(percentage_change_file, index=False, engine="openpyxl")

                    # Open the output file and update with additional data
                    workbook = load_workbook(output_file)
                    sheet = workbook.active

                    # Write the average price to row 2, column 10
                    sheet.cell(row=2, column=9, value="Average price (USD)")
                    sheet.cell(row=2, column=10, value=average_price_of_50_cryptos)

                    workbook.save(output_file)

                    # Display the results in the console
                    print(f"Maximum percentage change in 24 hours: {percentage_change_sorted.iloc[0]['Name']} = {percentage_change_sorted.iloc[0]['24h % Change']} %")
                    print(f"Minimum percentage change in 24 hours: {percentage_change_sorted.iloc[49]['Name']} = {percentage_change_sorted.iloc[49]['24h % Change']} %")
                    print(f"Average price of the 50 cryptocurrencies is {average_price_of_50_cryptos} USD")

                    workbook = load_workbook(percentage_change_file)
                    sheet = workbook.active

                    # Write the maximum and minimum percentage change
                    sheet.cell(row=2, column=9, value="Maximum change")
                    sheet.cell(row=2, column=10, value=percentage_change_sorted.iloc[0]["Name"])
                    sheet.cell(row=2, column=11, value="Minimum change")
                    sheet.cell(row=2, column=12, value=percentage_change_sorted.iloc[49]["Name"])

                    workbook.save(percentage_change_file)

                    # Authenticate and upload to Google Drive
                    drive_service = authenticate_google_drive(service_account_json)
                    upload_or_update_file(output_file, drive_service, parent_folder)
                    upload_or_update_file(percentage_change_file, drive_service, parent_folder)

                    print(f"Updated at: {time.ctime()}")
                    print("=" * 50)

                except PermissionError as e:
                    print(f"File writing error: {e}")

        # Wait for 5 minutes before fetching data again
        time.sleep(300)

except KeyboardInterrupt:
    print("Program stopped by user.")
